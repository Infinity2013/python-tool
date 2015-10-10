#!/usr/bin/env python
import openpyxl
import re
import sys
import os
import operator
from pprint import pprint

pos_dict = [ 
        ['c6', 'd6'],
        ['c7', 'd7'],
        ['c8', 'd8'],
        ['c9', 'd9'],
        ['c10', 'd10'],
        ['c13', 'd13'],
        ['c14', 'd14'],
        ['c15', 'd15'],
        ['c16', 'd16'],
        ['c17', 'd17'],
        ['c20', 'd20'],
        ['c23', 'd23'],
        ['c24', 'd24'],
        ['c25', 'd25'],
        ['c28', 'd28'],
        ['c31', 'd31'],
        ['c32', 'd32'],
        ['c33', 'd33'],
        ['c36', 'd36'],
        ['c37', 'd37'],
        ['c40', 'd40'],
        ['c43', 'd43'],
        ['c46', 'd46'],
        ['c47', 'd47'],
        ['c48', 'd48'],
        ['c49', 'd49'],
        ['c50', 'd50'],
        ['c51', 'd51'],
        ['c52', 'd52'],
        ['c55', 'd55'],
        ['c56', 'd56'],
        ['c57', 'd57'],
        ['c60', 'd60'],
]
def parse_final(f_name):
    with open("%s/%s" % (os.getcwd(), f_name),  "r") as f:
        t = f.readlines()
        last_key = ""
        res_list = [] 
        index = 0
        for line in t:
            g = re.match(r'real\s+(?P<min>\d+)m\s+(?P<sec>[\d\.]+)s.*', line)
            if g is not None:
                res_list.append([g.group("min"), g.group("sec")])
        return res_list

def write_to_xlsx(f_xlsx, res_list, pos_dict):
    wb = openpyxl.load_workbook("%s/%s" % (os.getcwd(), f_xlsx))
    ws = wb.active
    step = 1 if len(res_list) == len(pos_dict) else 4
    for i in xrange(0, len(res_list), step):
        t = res_list[i: i + step]
        sorted(t, key=operator.itemgetter(0, 1))
        pprint(t)
        t = t[0]
        for j in xrange(len(t)):
            ws[pos_dict[i / step][j]].value = t[j]

    wb.save("%s/%s" % (os.getcwd(), "new.xlsx"))

def get_res():
    wb = openpyxl.load_workbook("%s/%s" % (os.getcwd(), "new.xlsx"))
    ws = wb.active

def main():
    res_list = parse_final(sys.argv[1])

    write_to_xlsx(sys.argv[2], res_list, pos_dict)

if __name__ == '__main__':
    main()
